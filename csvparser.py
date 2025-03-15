import pandas as pd
from datetime import datetime
import pytz
from openpyxl import load_workbook

def auto_adjust_column_width(excel_file):
    """ Adjust column widths dynamically based on content length """
    wb = load_workbook(excel_file)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # Get column letter (e.g., 'A', 'B')
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2  # Add some padding
    wb.save(excel_file)

def convert_timezone(df, timestamp_cols, tz_from, tz_to):
    """ Convert all timestamp columns from one timezone to another """
    tz_from = pytz.timezone(tz_from)
    tz_to = pytz.timezone(tz_to)

    for col in timestamp_cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')  # Convert to datetime
            df[col] = df[col].dt.tz_localize(tz_from, ambiguous='NaT').dt.tz_convert(tz_to)  
            df[col] = df[col].dt.strftime("%Y-%m-%d %H:%M:%S %z")  # Keep timezone offset as string
    
    return df

def merge_csvs(csv1, csv2, csv3, output_file):
    # Load CSVs
    df1 = pd.read_csv(csv1, usecols=["RunTime", "ExecutableName"])
    df2 = pd.read_csv(csv2, usecols=["BagPath", "AbsolutePath", "CreatedOn", "ModifiedOn", "AccessedOn", "LastWriteTime", "FirstInteracted", "LastInteracted"])
    df3 = pd.read_csv(csv3, usecols=["SourceFile", "SourceCreated", "SourceModified", "SourceAccessed", "LocalPath", "MachineMACAddress"])

    # Define a unified column structure
    unified_columns = [
        "AccessTime", "Executable/Path", 
        "CreatedOn", "ModifiedOn", "LastWriteTime", "FirstInteracted", "LastInteracted",
        "BagPath", "LocalPath", "MachineMACAddress", "SourceFile"
    ]

    # Normalize column names in each DataFrame
    df1 = df1.rename(columns={"RunTime": "AccessTime", "ExecutableName": "Executable/Path"})
    df2 = df2.rename(columns={"CreatedOn": "CreatedOn", "ModifiedOn": "ModifiedOn", "AccessedOn": "AccessTime", "LastWriteTime": "LastWriteTime",
                               "FirstInteracted": "FirstInteracted", "LastInteracted": "LastInteracted", "AbsolutePath": "Executable/Path"})
    df3 = df3.rename(columns={"SourceFile": "SourceFile", "SourceCreated": "CreatedOn", "SourceModified": "ModifiedOn", "SourceAccessed": "AccessTime",
                               "LocalPath": "Executable/Path", "MachineMACAddress": "MachineMACAddress"})

    # Fill missing columns with NaN
    for df in [df1, df2, df3]:
        for col in unified_columns:
            if col not in df.columns:
                df[col] = None

    # Merge all data
    merged_df = pd.concat([df1, df2, df3], ignore_index=True)

    # Sort by AccessTime, then LastWriteTime if AccessTime is missing
    merged_df["AccessTime"] = pd.to_datetime(merged_df["AccessTime"], errors='coerce')
    merged_df["LastWriteTime"] = pd.to_datetime(merged_df["LastWriteTime"], errors='coerce')
    merged_df = merged_df.sort_values(by=["AccessTime", "LastWriteTime"], ascending=True)

    # Define all timestamp columns
    timestamp_cols = ["AccessTime", "CreatedOn", "ModifiedOn", "LastWriteTime", "FirstInteracted", "LastInteracted"]

    # Convert timezones AFTER sorting
    merged_df_utc = convert_timezone(merged_df.copy(), timestamp_cols, "UTC", "UTC")
    merged_df_utc8 = convert_timezone(merged_df.copy(), timestamp_cols, "UTC", "Asia/Singapore")

    # Save to an Excel file with two sheets (UTC+0000 and UTC+0800)
    with pd.ExcelWriter(output_file) as writer:
        merged_df_utc.to_excel(writer, sheet_name="UTC+0000", index=False)
        merged_df_utc8.to_excel(writer, sheet_name="UTC+0800", index=False)

    # Adjust column widths in the Excel file
    auto_adjust_column_width(output_file)

    print(f"Final timeline saved to {output_file} with two timezone sheets, and columns auto-adjusted.")

# Example usage
merge_csvs("prefetch_output_Timeline.csv", "John_UsrClass_shellb_output.csv", "link_output.csv", "final_output.xlsx")
